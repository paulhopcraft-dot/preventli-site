"""
Decode step 2: dump the engine chain (formula | cached value), no recalc.
- All defined names resolved to formula + value
- Full EPR sheet grid
- Z sheet computational region
- Data input region (cached Princes inputs)
- Key Parameters
"""
import openpyxl, re
from openpyxl.utils import range_boundaries, get_column_letter

SRC = r"D:\Downloads\Premium Master Cracked 201213 Simulator (1).xlsx"
wb_f = openpyxl.load_workbook(SRC, data_only=False)
wb_v = openpyxl.load_workbook(SRC, data_only=True)

def cell(sheet, coord):
    f = wb_f[sheet][coord].value
    v = wb_v[sheet][coord].value
    return f, v

def fmt(x):
    if isinstance(x, float):
        return f"{x:.6f}".rstrip('0').rstrip('.') if x == x else str(x)
    return repr(x)

print("=" * 90)
print("NAMED RANGES  (name = formula | value)")
print("=" * 90)
dn = wb_f.defined_names
for name in sorted(dn.keys()):
    try:
        ref = dn[name].value  # e.g. EPR!$P$15  or  'Empr ThresholdX'!$V$7
        # single-cell only
        m = re.match(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)$", ref)
        if not m:
            print(f"  {name:22} -> {ref}  (range/complex)")
            continue
        sh, col, row = m.group(1), m.group(2), m.group(3)
        coord = f"{col}{row}"
        if sh not in wb_f.sheetnames:
            print(f"  {name:22} -> {ref}  (sheet missing)")
            continue
        f, v = cell(sh, coord)
        print(f"  {name:22} {sh}!{coord:5} = {fmt(v):>16}   <- {f!r}")
    except Exception as e:
        print(f"  {name:22} -> <err {e}>")

print()
print("=" * 90)
print("EPR SHEET  (full grid: coord = value  <- formula)  rows 2-44")
print("=" * 90)
wsf, wsv = wb_f["EPR"], wb_v["EPR"]
for row in range(2, 45):
    for col in range(1, 18):  # A..Q
        L = get_column_letter(col)
        coord = f"{L}{row}"
        f = wsf[coord].value
        v = wsv[coord].value
        if f is None and v is None:
            continue
        tag = "F" if (isinstance(f, str) and f.startswith("=")) else " "
        print(f"  {coord:5} [{tag}] = {fmt(v):>16}   <- {f!r}")

print()
print("=" * 90)
print("Z SHEET  rows 1-14, cols A-L  (value <- formula)")
print("=" * 90)
wsf, wsv = wb_f["Z"], wb_v["Z"]
for row in range(1, 15):
    for col in range(1, 13):
        L = get_column_letter(col)
        coord = f"{L}{row}"
        f = wsf[coord].value
        v = wsv[coord].value
        if f is None and v is None:
            continue
        print(f"  Z!{coord:5} = {fmt(v):>16}   <- {f!r}")
