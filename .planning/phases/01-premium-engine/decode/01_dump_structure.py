"""
Decode step 1: dump the 2012-13 cracked simulator structure.
- Sheet names
- All defined names (named ranges) + their refs
- Key cells from the handoff: EPR!B12,B18,E15,P15 ; Z!D8,H9 — formula + cached value
No recalc. We read what Excel already cached.
"""
import openpyxl
from openpyxl.utils import get_column_letter

SRC = r"D:\Downloads\Premium Master Cracked 201213 Simulator (1).xlsx"

# data_only=False -> formulas ; data_only=True -> last cached values
wb_f = openpyxl.load_workbook(SRC, data_only=False, read_only=False)
wb_v = openpyxl.load_workbook(SRC, data_only=True, read_only=False)

print("=" * 70)
print("SHEETS")
print("=" * 70)
for i, ws in enumerate(wb_f.worksheets):
    dims = ws.dimensions
    print(f"  [{i:2}] {ws.title!r:40} state={ws.sheet_state:8} dims={dims}")

print()
print("=" * 70)
print("DEFINED NAMES (named ranges)")
print("=" * 70)
dn = wb_f.defined_names
try:
    names = list(dn.keys())
except Exception:
    names = [d.name for d in dn.definedName]
for name in sorted(names):
    try:
        d = dn[name]
        print(f"  {name:30} -> {d.value}")
    except Exception as e:
        print(f"  {name:30} -> <err {e}>")

print()
print("=" * 70)
print("KEY CELLS (formula | cached value)")
print("=" * 70)
targets = [
    ("EPR", ["B12", "B18", "E15", "P15", "E12", "E18"]),
    ("Z",   ["D8", "H9", "D9", "H8"]),
]
for sheet, cells in targets:
    if sheet not in wb_f.sheetnames:
        print(f"  !! sheet {sheet!r} not found")
        continue
    wsf = wb_f[sheet]
    wsv = wb_v[sheet]
    print(f"-- sheet {sheet!r} --")
    for c in cells:
        f = wsf[c].value
        v = wsv[c].value
        print(f"   {sheet}!{c:5} formula={f!r:55} value={v!r}")
