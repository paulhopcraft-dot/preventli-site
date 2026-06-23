"""
Decode step 3: dump EPR grid, Z grid, both Preliminary Prem sheets, and key
Parameters to a UTF-8 file (avoids console encoding crash). No recalc.
"""
import openpyxl
from openpyxl.utils import get_column_letter

SRC = r"D:\Downloads\Premium Master Cracked 201213 Simulator (1).xlsx"
OUT = r"D:\dev\preventli-site\.planning\phases\01-premium-engine\decode\grids.txt"

wb_f = openpyxl.load_workbook(SRC, data_only=False)
wb_v = openpyxl.load_workbook(SRC, data_only=True)

def fmt(x):
    if isinstance(x, float):
        return f"{x:.6f}"
    s = str(x)
    return s if len(s) <= 40 else s[:37] + "..."

lines = []
def dump_grid(sheet, r0, r1, c0, c1, label):
    lines.append("=" * 90)
    lines.append(f"{label}  ({sheet}!{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{r1})")
    lines.append("=" * 90)
    wsf, wsv = wb_f[sheet], wb_v[sheet]
    for row in range(r0, r1 + 1):
        for col in range(c0, c1 + 1):
            L = get_column_letter(col)
            coord = f"{L}{row}"
            f = wsf[coord].value
            v = wsv[coord].value
            if f is None and v is None:
                continue
            tag = "F" if (isinstance(f, str) and f.startswith("=")) else " "
            lines.append(f"  {sheet}!{coord:5} [{tag}] = {fmt(v):>18}   <- {f!r}")
    lines.append("")

dump_grid("EPR", 2, 44, 1, 17, "EPR SHEET")            # A..Q
dump_grid("Z", 1, 16, 1, 18, "Z SHEET")                # A..R
dump_grid("2012-13 Preliminary Prem", 1, 30, 1, 20, "2012-13 PRELIM PREM")   # A..T
dump_grid("2011-12 Preliminary Prem", 1, 30, 1, 20, "2011-12 PRELIM PREM")
dump_grid("Parameters", 2, 30, 4, 10, "PARAMETERS D..J rows2-30")

with open(OUT, "w", encoding="utf-8") as fh:
    fh.write("\n".join(lines))
print(f"wrote {len(lines)} lines to {OUT}")
